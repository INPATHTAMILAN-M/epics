from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
import random
from django.core.paginator import Paginator
from .forms import StudentForm
from .models import Student, Problem_statement
from django.shortcuts import get_object_or_404
from django.conf import settings

from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404

from django.db.models import Count
from django.core.mail import send_mail
from django.http import JsonResponse
import random

def index(request):
    form = StudentForm()

    # Filters
    title_query = request.GET.get('title', '')
    theme_query = request.GET.get('theme', '')
    category_query = request.GET.get('category', '')
    problems = Problem_statement.objects.select_related('theme', 'created_by')
    count = Problem_statement.objects.annotate(
    submission_count=Count('student')
)
    
    if title_query:
        problems = problems.filter(title__icontains=title_query)
    if theme_query:
        problems = problems.filter(theme__theme__iexact=theme_query)
    if category_query:
        problems = problems.filter(category__iexact=category_query)

    # Pagination
    paginator = Paginator(problems, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'count': count,
        'form': form,
        'page_obj': page_obj,
        'title_query': title_query,
        'theme_query': theme_query,
        'category_query': category_query,
        'themes': Problem_statement.objects.values_list('theme__theme', flat=True).distinct(),
        'categories': Problem_statement.objects.values_list('category', flat=True).distinct(),
        'total_problems': Problem_statement.objects.count(),
        'hardware_count': Problem_statement.objects.filter(category='hardware').count(),
        'software_count': Problem_statement.objects.filter(category='software').count(),
        'total_submissions': Student.objects.count(),
    }
    return render(request, 'index.html', context)


def apply_problem_view(request, problem_id):
    problem = get_object_or_404(Problem_statement, pk=problem_id)
    
    form = StudentForm()

    if request.method == 'POST':
        print(problem)
        form = StudentForm(request.POST, request.FILES)
        
        print("before if")
        if form.is_valid():
            print("Hii")
            student = form.save(commit=False)
            student.problem = problem  # Assign the problem to the student
            student.save()
            messages.success(request, "Application submitted successfully!")
            return redirect('index')  # Redirect after successful submission
        else:
            print(form.errors)  # Print form errors for debugging
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('apply_form_partial.html', {'form': form, 'problem': problem}, request)
        return HttpResponse(html)

    return render(request, 'apply_form_partial.html', {'form': form, 'problem': problem})


@csrf_exempt
def send_otp_view(request):
    if request.method == 'POST':


        email = request.POST.get('email')

        if '@karpagamtech.ac.in' not in email:
            return JsonResponse({'success': False, 'message': 'Use college email to submit solution'}, status=400)
        
        teamleader_name = request.POST.get('teamleader_name')
        if not email:
            return JsonResponse({'success': False, 'message': 'Email is required'}, status=400)

        otp = str(random.randint(100000, 999999))

        

        # Send email
        subject = 'Your OTP for EPICS Application'
        message = f'Hello {teamleader_name}, \n Your OTP is: {otp}'
        from_email = 'skssoftwaretech@gmail.com'
        recipient_list = [email]

        try:
            send_mail(subject, message, from_email, recipient_list)
        except Exception as e:
            return JsonResponse({'success': False, 'message': 'Failed to send email', 'error': str(e)}, status=500)

        return JsonResponse({'success': True, 'otp': otp})
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=405)


# @csrf_exempt
# def verify_otp(request):
#     if request.method == 'POST':
#         entered_otp = request.POST.get('otp')
#         saved_otp = request.session.get('otp')
#         if entered_otp == saved_otp:
#             request.session['otp_verified'] = True
#             return JsonResponse({'status': 'success'})
#         return JsonResponse({'status': 'fail'})
#     return JsonResponse({'status': 'error'})

# def submit_application(request):
#     if request.method == 'POST':
#         if request.session.get('otp_verified'):
#             form = StudentForm(request.POST, request.FILES)
#             if form.is_valid():
#                 form.save()
#                 messages.success(request, "Application submitted successfully!")
#                 del request.session['otp']
#                 del request.session['otp_verified']
#                 return JsonResponse({'success': False, 'message': 'Use college email to submit solution'}, status=400)
#                 return redirect('index')
#             else:
#                 messages.error(request, "Form invalid. Please check the fields.")
#         else:
#             messages.error(request, "OTP verification required before submitting.")
#         return redirect('index')

def guidelines(request):
    return render(request, 'guidelines.html')

def timeline(request):
    return render(request, 'timeline.html')



import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import ExcelUploadForm
from .models import Problem_statement, Theme, Faculty
from openpyxl import load_workbook
from django.conf import settings

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                theme_id, problem_id, title, description, category, created_by_id = row
                
                # Check if the theme_id and created_by_id are valid integers
                try:
                    theme = Theme.objects.get(id=int(theme_id))
                    created_by = Faculty.objects.get(id=int(created_by_id))

                    # Create a new Problem Statement instance
                    problem_statement = Problem_statement(
                        theme=theme,
                        problem_id=problem_id,
                        title=title,
                        description=description,
                        category=category,
                        created_by=created_by
                    )
                    problem_statement.save()

                except ValueError:
                    messages.error(request, f"Invalid ID in row {row}. Please ensure theme_id and created_by_id are valid numbers.")
                    continue
                except Theme.DoesNotExist:
                    messages.error(request, f"Theme with ID {theme_id} does not exist in row {row}.")
                    continue
                except Faculty.DoesNotExist:
                    messages.error(request, f"Faculty with ID {created_by_id} does not exist in row {row}.")
                    continue
            
            messages.success(request, 'Excel file successfully uploaded!')
            return redirect('upload_excel')  # Redirect after success

    else:
        form = ExcelUploadForm()

    return render(request, 'upload.html', {'form': form})


